

#!/usr/bin/env python
# -*- coding: utf-8 -*-
# region Description
__author__ = 'TruongNV - NGUYEN VAN TRUONG'
__copyright__ = "Copyright ©2023 TruongNV <truongg.nv@gmail.com>"
__maintainer__ = "TruongNV"
__email__ = "truongg.nv@gmail.com"
__status__ = "Production"
__date__ = 24/8/2023
# endregion

from django.http import HttpResponse
from django.template import loader
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from SystemManagement.tnv_decorator import org_require as org_login_required
# from django.contrib.gis.db.models import Union
from django.shortcuts import render
from django.urls import URLPattern, URLResolver
from django.shortcuts import render, get_object_or_404, redirect
from django import template
from django.conf import settings
from django.http import JsonResponse
from django.http import Http404
from datetime import datetime as dtime
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.db.models import Q
from Workspace.views import AdminWebContext
from uuid import uuid4 as UUID4
import uuid
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import sys
import os
import json
from .models import *


@org_login_required(login_url="/login/")
def Index(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:Index', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    context["pagename"] = "SentMailManagement"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name":"Dashboard Quản lý email",
            "link": context['org_app_prefix'] + "/SentMailManagement/",
        },
        
    ]
    template = loader.get_template(str('SentMailManagement/%s/admin/index.html' % context['website_template']))
    
    context['systemmailsenderconfig_count'] = SystemMailSenderConfig.objects.all().count()

                
    context['systemmailtarget_count'] = SystemMailTarget.objects.all().count()

                
    context['systemmailmail_count'] = SystemMailMail.objects.all().count()

                
    context['systemmailsentmail_count'] = SystemMailSentMail.objects.all().count()

                
    context['systemmailserverconfig_count'] = SystemMailServerConfig.objects.all().count()

                
    context['systemmailtemplate_count'] = SystemMailTemplate.objects.all().count()

                
    context['systemmailtype_count'] = SystemMailType.objects.all().count()

                
    context['systemmailtemplateimportdata_count'] = SystemMailTemplateImportData.objects.all().count()

                
    context['systemmailtemplateimportstatus_count'] = SystemMailTemplateImportStatus.objects.all().count()

                
    return HttpResponse(template.render(context, request))

from .RequestHandler.TNVSystemMailSenderConfigRequestHandler import systemmailsenderconfig_post_handle, systemmailsenderconfig_get_handle

@org_login_required(login_url="/login/")
def SystemMailSenderConfigView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSenderConfig_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfig_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSenderConfigCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSenderConfig/SystemMailSenderConfig-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsenderconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email gửi đi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email gửi đi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSenderConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsenderconfig'] = SystemMailSenderConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSenderConfig_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailSenderConfig.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailSenderConfig_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailSenderConfig.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailTargetRequestHandler import systemmailtarget_post_handle, systemmailtarget_get_handle

@org_login_required(login_url="/login/")
def SystemMailTargetView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTarget_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTarget_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTarget_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTarget_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTarget_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTarget_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTarget_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTarget_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTarget_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTargetCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTarget/SystemMailTarget-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtarget_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cá nhân/nhóm nhận mail"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cá nhân/nhóm nhận mail",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTarget/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtarget'] = SystemMailTarget.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTarget_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTarget.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailTarget_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTarget.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailMailRequestHandler import systemmailmail_post_handle, systemmailmail_get_handle

@org_login_required(login_url="/login/")
def SystemMailMailView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailMail_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailMail_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailMail_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailMail_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailMail_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailMail_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailMail_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMail_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailMail_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailMailCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailMail/SystemMailMail-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail cần gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail cần gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailmail'] = SystemMailMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailMail_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailMail.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailMail_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailMail.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailSentMailRequestHandler import systemmailsentmail_post_handle, systemmailsentmail_get_handle

@org_login_required(login_url="/login/")
def SystemMailSentMailView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSentMail_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSentMail_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSentMail_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSentMail_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSentMail_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSentMail_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailSentMail_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMail_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailSentMail_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailSentMailCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailSentMail/SystemMailSentMail-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailsentmail_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mail đã gửi"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mail đã gửi",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailSentMail/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailsentmail'] = SystemMailSentMail.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailSentMail_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailSentMail.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailSentMail_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailSentMail.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailServerConfigRequestHandler import systemmailserverconfig_post_handle, systemmailserverconfig_get_handle

@org_login_required(login_url="/login/")
def SystemMailServerConfigView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailServerConfig_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailServerConfig_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailServerConfig_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailServerConfig_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailServerConfig_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfig_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailServerConfig_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailServerConfigCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailServerConfig/SystemMailServerConfig-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailserverconfig_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Cấu hình email server"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Cấu hình email server",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailServerConfig/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailserverconfig'] = SystemMailServerConfig.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailServerConfig_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailServerConfig.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailServerConfig_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailServerConfig.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailTemplateRequestHandler import systemmailtemplate_post_handle, systemmailtemplate_get_handle

           
@org_login_required(login_url="/login/")
def SystemMailTemplate_ImportFileView(request, *args, **kwargs):
    crr_status = 500
    response = {}
    response["status"] = "fail"
    if request.method == 'POST':
        if len(request.FILES) > 0:
            crr_import_file = request.FILES["file"]
            # Lưu file 
            import_data_systemmailtemplate_obj = SystemMailTemplateImportData()
            import_data_systemmailtemplate_obj.file = crr_import_file
            import_data_systemmailtemplate_obj.name = crr_import_file.name
            # import_data_systemmailtemplate_obj.created_by = request.user
            import_data_systemmailtemplate_obj.save()

            # list data import 
            import_raw_data_list = []
            import_obj_list = []
            # list wrong or fail data
            fail_result_list = []
            validation_result_list = []
            wb_obj = openpyxl.load_workbook(crr_import_file, data_only=True)
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            print(f'max_col = {max_col}')
            max_row = sheet_obj.max_row
            print(f'max_row = {max_row}')
            start_point = [1, 4]
            header_row = 3 # dòng của header
            start_col = start_point[0]
            start_row = start_point[1]
            list_col_name = []
            for i in range(start_col, max_col + 1):
                # Loop will print all columns name
                crr_col_name = str(sheet_obj.cell(row=header_row, column=i).value).strip()
                if crr_col_name != "" and crr_col_name != None and crr_col_name != "None":
                    if hasattr(SystemMailTemplate(), crr_col_name):
                        list_col_name.append({
                            'name': crr_col_name,
                            'col_number': i,
                        })
            if len(list_col_name) > 0 :
                for y in range(start_row, max_row + 1):
                    new_import_data = {}
                    new_import_obj = {}
                    validate_field_import_obj = {}
                    fail_result_import_obj = {}
                    try:
                        # Loop will get value row by row
                        for col in list_col_name:
                        # Loop will print all columns name
                            crr_col_value = str(sheet_obj.cell(row=y, column=col["col_number"]).value).strip()
                            crr_col_value_validated = crr_col_value
                            if crr_col_value != "" and crr_col_value != None and crr_col_value != "None":
                                #validate
                                
                                # validate giá trị Thời điểm cập nhật(DateTimeField)
                                try:
                                    if col["name"] == "updated_at":
                                        crr_col_value_validated = datetime.strptime(str(crr_col_value), '%d/%m/%Y %H:%M')
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm cập nhật",
                                            "field_name":"updated_at",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                            "mess":"Không đúng định dạng ('%d/%m/%Y %H:%M')(vd: '30/04/1975 23:30')",
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Loại email(ForeignKey)
                                try:
                                    

                                    if col["name"] == "email_type":
                                        crr_uuid = None
                                        try:
                                            crr_uuid = uuid.UUID(str(crr_col_value))
                                        except Exception as xx:
                                            print("email_type not is uuid")
                                        
                                        if crr_uuid != None:
                                            crr_obj_relation = SystemMailType.objects.filter(uuid=crr_uuid).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Loại email",
                                                        "field_name":"email_type",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                                
                                        else:
                                            crr_obj_relation = SystemMailType.objects.filter(Q(friendly_code=crr_col_value) | Q(code=crr_col_value)).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Loại email",
                                                        "field_name":"email_type",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Loại email",
                                            "field_name":"email_type",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Mặc định(BooleanField)
                                try:
                                    if col["name"] == "is_default":
                                        crr_col_value_validated = float(str(crr_col_value))
                                        if crr_col_value.lower() in ["true","false","1","0"]:
                                            if crr_col_value.lower() in ["true","1"]:
                                                crr_col_value_validated = True
                                            else:
                                                crr_col_value_validated = False
                                        else:
                                            validation_result_list.append(
                                            {
                                                "field_verbose_name":"Mặc định",
                                                "field_name":"is_default",
                                                "field_value":crr_col_value,
                                                "mess":"Không đúng định dạng",
                                            }
                                            )
                                        continue
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Mặc định",
                                            "field_name":"is_default",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                            "mess":"Không đúng định dạng",
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Tổ chức(ForeignKey)
                                try:
                                    

                                    if col["name"] == "organization":
                                        crr_uuid = None
                                        try:
                                            crr_uuid = uuid.UUID(str(crr_col_value))
                                        except Exception as xx:
                                            print("organization not is uuid")
                                        
                                        if crr_uuid != None:
                                            crr_obj_relation = Organization.objects.filter(uuid=crr_uuid).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Tổ chức",
                                                        "field_name":"organization",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                                
                                        else:
                                            crr_obj_relation = Organization.objects.filter(Q(friendly_code=crr_col_value) | Q(code=crr_col_value)).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Tổ chức",
                                                        "field_name":"organization",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Tổ chức",
                                            "field_name":"organization",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Người tạo(ForeignKey)
                                try:
                                    

                                    if col["name"] == "hr_created_by":
                                        crr_uuid = None
                                        try:
                                            crr_uuid = uuid.UUID(str(crr_col_value))
                                        except Exception as xx:
                                            print("hr_created_by not is uuid")
                                        
                                        if crr_uuid != None:
                                            crr_obj_relation = HRAccount.objects.filter(uuid=crr_uuid).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Người tạo",
                                                        "field_name":"hr_created_by",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                                
                                        else:
                                            crr_obj_relation = HRAccount.objects.filter(Q(friendly_code=crr_col_value) | Q(code=crr_col_value)).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Người tạo",
                                                        "field_name":"hr_created_by",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người tạo",
                                            "field_name":"hr_created_by",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Người cập nhật(ForeignKey)
                                try:
                                    

                                    if col["name"] == "hr_updated_by":
                                        crr_uuid = None
                                        try:
                                            crr_uuid = uuid.UUID(str(crr_col_value))
                                        except Exception as xx:
                                            print("hr_updated_by not is uuid")
                                        
                                        if crr_uuid != None:
                                            crr_obj_relation = HRAccount.objects.filter(uuid=crr_uuid).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Người cập nhật",
                                                        "field_name":"hr_updated_by",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                                
                                        else:
                                            crr_obj_relation = HRAccount.objects.filter(Q(friendly_code=crr_col_value) | Q(code=crr_col_value)).all()
                                            if len(crr_obj_relation) == 1:
                                                crr_col_value_validated = crr_obj_relation.values("uuid","code","name").first()
                                            elif len(crr_obj_relation) > 1:
                                                validation_result_list.append(
                                                    {
                                                        "field_verbose_name":"Người cập nhật",
                                                        "field_name":"hr_updated_by",
                                                        "field_value":crr_col_value,
                                                        "mess":f"Có nhiều hơn 1 đối tượng dành cho giá trị của bạn {crr_col_value}",
                                                    }
                                                )
                                                continue
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người cập nhật",
                                            "field_name":"hr_updated_by",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                        }
                                    )
                                    continue
                        
                                # validate giá trị Thời điểm tạo(DateTimeField)
                                try:
                                    if col["name"] == "created_at":
                                        crr_col_value_validated = datetime.strptime(str(crr_col_value), '%d/%m/%Y %H:%M')
                                except Exception as xx:
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm tạo",
                                            "field_name":"created_at",
                                            "field_value":crr_col_value,
                                            "sys_mess":str(xx),
                                            "mess":"Không đúng định dạng ('%d/%m/%Y %H:%M')(vd: '30/04/1975 23:30')",
                                        }
                                    )
                                    continue
                        
                                # add value to attr
                                
                                new_import_data[col["name"]] = crr_col_value
                                new_import_obj[col["name"]] = crr_col_value_validated
                                pass
                            else:
                                # kiểm tra bắt buộc điền
                                
                                if crr_col_value == "" and col["name"] == "name":
                                    # Tên bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Tên",
                                            "field_name":"name",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "name":
                                    # Tên bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Tên",
                                            "field_name":"name",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "code":
                                    # Mã bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Mã",
                                            "field_name":"code",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "code":
                                    # Mã bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Mã",
                                            "field_name":"code",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "friendly_code":
                                    # Mã bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Mã",
                                            "field_name":"friendly_code",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "created_by":
                                    # Người tạo bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người tạo",
                                            "field_name":"created_by",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "created_by":
                                    # Người tạo bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người tạo",
                                            "field_name":"created_by",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "updated_by":
                                    # Người cập nhật bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người cập nhật",
                                            "field_name":"updated_by",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "updated_by":
                                    # Người cập nhật bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Người cập nhật",
                                            "field_name":"updated_by",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "updated_at":
                                    # Thời điểm cập nhật bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm cập nhật",
                                            "field_name":"updated_at",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "updated_at":
                                    # Thời điểm cập nhật bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm cập nhật",
                                            "field_name":"updated_at",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                if crr_col_value == "" and col["name"] == "created_at":
                                    # Thời điểm tạo bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm tạo",
                                            "field_name":"created_at",
                                            "mess":"Bắt buộc điền",
                                        }
                                    )
                                    continue
                            
                                if (crr_col_value == None or crr_col_value == "None") and col["name"] == "created_at":
                                    # Thời điểm tạo bắt buộc điền
                                    validation_result_list.append(
                                        {
                                            "field_verbose_name":"Thời điểm tạo",
                                            "field_name":"created_at",
                                            "mess":"Không được phép None(null)",
                                        }
                                    )
                                    continue
                            
                                pass
                    except Exception as xx:
                        print(str(xx))
                        
                        fail_result_list.append({
                            "row_num": y,
                            "err_mess": str(xx),
                        })
                    print(new_import_data)
                    import_raw_data_list.append(new_import_data)
                    import_obj_list.append(new_import_obj)
                import_data_systemmailtemplate_obj.json_import_data = json.dumps(import_raw_data_list)
                import_data_systemmailtemplate_obj.json_check_validation = json.dumps(validation_result_list)
                import_data_systemmailtemplate_obj.json_fail_data = json.dumps(fail_result_list)
                if len(validation_result_list) > 0 or len(fail_result_list) > 0:
                    import_data_systemmailtemplate_obj.is_import_failed = True
                    import_data_systemmailtemplate_obj.is_import_success = False
                    response["status"] = "import-err"
                    crr_status = 500
                else:
                    for idata in import_obj_list:
                        SystemMailTemplate.objects.create(**idata)
                    import_data_systemmailtemplate_obj.is_import_failed = False
                    import_data_systemmailtemplate_obj.is_import_success = True
                    response["status"] = "ok"
                    crr_status = 200
                import_data_systemmailtemplate_obj.save()
                # trả về kết quả import data
                response["is_import_failed"] = import_data_systemmailtemplate_obj.is_import_failed 
                response["is_import_success"] = import_data_systemmailtemplate_obj.is_import_success 
                response["fail_result_list"] = fail_result_list
                response["validation_result_list"] = validation_result_list
    return JsonResponse(response, status=crr_status)

@org_login_required(login_url="/login/")
def SystemMailTemplate_ImportFileTemplateView(request, *args, **kwargs):
    crr_status = 500
    response = {}
    response["status"] = "fail"
    
    #Making this active worksheet
    Wb_test = openpyxl.Workbook()
    sheet = Wb_test.active
    
    sheet.title = "Mẫu email"
    
    # create content file 

    start_point = [1,2]
    start_col = start_point[0]
    start_row = start_point[1]
    all_field = [
        
                        {
                                "name":"friendly_code",
                                "verbose_name":"Mã",
                        },
                        
                        {
                                "name":"name",
                                "verbose_name":"Tên",
                        },
                        
                        {
                                "name":"email_type",
                                "verbose_name":"Loại email",
                        },
                        
                        {
                                "name":"content",
                                "verbose_name":"Nội dung",
                        },
                        
                        {
                                "name":"is_default",
                                "verbose_name":"Mặc định",
                        },
                        
                        {
                                "name":"organization",
                                "verbose_name":"Tổ chức",
                        },
                        
                        {
                                "name":"hr_created_by",
                                "verbose_name":"Người tạo",
                        },
                        
                        {
                                "name":"hr_updated_by",
                                "verbose_name":"Người cập nhật",
                        },
                        
    ]
    blue = "C5D9F1"
    count = start_col
    
    thin = Side(border_style="thin", color="000000")
    font = Font(bold=True)
    # double = Side(border_style="double", color="000000")
    for field in all_field: 
        c1 = sheet.cell(row = start_row, column = count)
        c2 = sheet.cell(row = start_row + 1, column = count)
        c1.value = str(field["verbose_name"])
        c2.value = str(field["name"])
        # color blue
        c1.fill = PatternFill(start_color=blue, end_color=blue,
                                        fill_type = "solid")
        c2.fill = PatternFill(start_color=blue, end_color=blue,
                                        fill_type = "solid")
        # border
        c1.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # text center
        c1.alignment = Alignment(horizontal='center')
        c2.alignment = Alignment(horizontal='center')
        # bold
        c1.font = font
        c2.font = font
        count += 1
    
    save_path = "SentMailManagement/excel/template/SystemMailTemplate-template.xlsx"
    #Creating second worksheet with its name
    os.chdir(settings.MEDIA_ROOT)
    if os.path.exists(os.path.dirname(save_path)) is False:
            os.system('mkdir -p %s' % os.path.dirname(save_path))
    
    Wb_test.save(save_path)
    if os.path.exists(save_path):
        with open(save_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(save_path)
            response["status"] = "ok"
            return response
    raise Http404

            


@org_login_required(login_url="/login/")
def SystemMailTemplateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplate_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplate_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplate_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplate_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplate_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplate_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplate_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplate_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplate_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplate/SystemMailTemplate-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplate_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplate/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplate'] = SystemMailTemplate.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplate_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplate.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailTemplate_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplate.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailTypeRequestHandler import systemmailtype_post_handle, systemmailtype_get_handle

@org_login_required(login_url="/login/")
def ComposeEmailView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:ComposeEmailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/ComposeEmail/ComposeEmail.html' % context['website_template']))
    # if request.method == 'POST':
    #     more_context = systemmailtype_post_handle(request)
    #     if more_context:
    #         context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Soạn email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Soạn email",
            "link": context['org_app_prefix'] + "/SentMailManagement/ComposeEmail/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result

@org_login_required(login_url="/login/")
def SystemMailTypeView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailType_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailType_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailType_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailType_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailType_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailType_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailType_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailType_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailType_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTypeCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailType/SystemMailType-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtype_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Loại email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Loại email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailType/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtype'] = SystemMailType.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailType_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailType.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailType_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailType.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailTemplateImportDataRequestHandler import systemmailtemplateimportdata_post_handle, systemmailtemplateimportdata_get_handle

@org_login_required(login_url="/login/")
def SystemMailTemplateImportDataView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportData_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportDataCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportData/SystemMailTemplateImportData-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportdata_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportData/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportdata'] = SystemMailTemplateImportData.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplateImportData.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportData_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplateImportData.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result

from .RequestHandler.TNVSystemMailTemplateImportStatusRequestHandler import systemmailtemplateimportstatus_post_handle, systemmailtemplateimportstatus_get_handle

@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatusView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_View', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QTVView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_QTVView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QLNSView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_QLNSView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_CTView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_CTView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)
    
    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Danh sách",
        },
    ]
    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_CreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Create.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QTVCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_QTVCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Create-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QLNSCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_QLNSCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Create-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_CTCreateView(request, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatus_CTCreateView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Create-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Thêm mới",
        },
    ]
    context['current_uuid'] = UUID4()

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_DetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Detail.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QTVDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusQTVDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Detail-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QLNSDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusQLNSDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Detail-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_CTDetailView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusCTDetailView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Detail-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Chi tiết",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_EditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Edit.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QTVEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusQTVEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Edit-QTV.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_QLNSEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusQLNSEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Edit-QLNS.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

            
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_CTEditView(request, slug, *args, **kwargs):
    try:
        context = AdminWebContext(request, page_info='SentMailManagement:SystemMailTemplateImportStatusCTEditView', *args, **kwargs)
    except Exception as xx:
        context = {}
        context['website_template'] = 'default'

    template = loader.get_template(str('SentMailManagement/%s/admin/SystemMailTemplateImportStatus/SystemMailTemplateImportStatus-Edit-CT.html' % context['website_template']))
    if request.method == 'POST':
        more_context = systemmailtemplateimportstatus_post_handle(request)
        if more_context:
            context.update(more_context)
    # Load danh sách bản ghi(nếu có)            
    

    context["pagename"] = "Trạng thái tải lên Mẫu email"
    context["breadcrumbs"] = [
        {
            "name": "Trang chủ",
            "link": context['org_prefix'] + "/overviews/",
        },
        {
            "name": context['this_application'].desc,
            "link": context["logo_direct_link"],
        },
        {
            "name":"Trạng thái tải lên Mẫu email",
            "link": context['org_app_prefix'] + "/SentMailManagement/SystemMailTemplateImportStatus/",
        },
        {
            "name":"Cập nhật",
        },
    ]
    try:
        context['systemmailtemplateimportstatus'] = SystemMailTemplateImportStatus.objects.filter(uuid=slug).first()
    except Exception as xx:
        print(str(xx))

    result = HttpResponse(template.render(context, request))
    return result
            

@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_AddingCheckExist(request, field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplateImportStatus.objects.filter(friendly_code=field_value).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
                                           
@org_login_required(login_url="/login/")
def SystemMailTemplateImportStatus_UpdateCheckExist(request, obj_uuid,field_value, *args, **kwargs):
    response = {}
    context = {}
    response["status"] = "exist"
    if field_value is None:
        response["status"] = "field-value-none"
        result = JsonResponse(response)
        return result                                 
    obj = SystemMailTemplateImportStatus.objects.filter(Q(friendly_code=field_value)&~Q(uuid=obj_uuid)).first()
    if obj != None:
        response["status_code"] = 403
        response["status"] = "exist"
    else:
        response["status_code"] = 200
        response["status"] = "ok"
    result = JsonResponse(response)
    return result
