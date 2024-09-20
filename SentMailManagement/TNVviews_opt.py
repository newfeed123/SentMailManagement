

#!/usr/bin/env python
# -*- coding: utf-8 -*-
# region Description
__author__ = 'TruongNV - NGUYEN VAN TRUONG'
__copyright__ = "Copyright ©2023 TruongNV <truongg.nv@gmail.com>"
__maintainer__ = "TruongNV"
__email__ = "truongg.nv@gmail.com"
__status__ = "Production"
__date__ = 24/8/2023
# endregion

from django.http import HttpResponse
from django.template import loader
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
# from django.contrib.gis.db.models import Union
from django.shortcuts import render
from django.urls import URLPattern, URLResolver
from django.shortcuts import render, get_object_or_404, redirect
from django import template
from django.conf import settings
from django.http import JsonResponse
from django.http import Http404
from datetime import datetime as dtime
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.db.models import Q
from Workspace.views import AdminWebContext
from uuid import uuid4 as UUID4
import uuid
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import sys
import os
import json
from .models import *


# dành cho việc cache data
from RedisCachingSystem.RedisModel import dictSearch,dictFilter
from core.settings import HRM_DOMAIN_NAME as DOMAIN_NAME
from django.core.paginator import Paginator
from django.http import JsonResponse
